import time
from seleniumbase import BaseCase
from selenium.webdriver.support.select import Select
from openpyxl import Workbook
from config.config import DataConfig
from selenium.common.exceptions import NoSuchElementException, TimeoutException

wb = Workbook()


class MyClass(BaseCase):

    def test_login_to_system(self):
        self.driver.maximize_window()
        self.driver.implicitly_wait(30)
        self.driver.set_page_load_timeout(30)

        try:
            self.open("https://app.cafcsn.it/#/login")
            self.wait_for_element("//input[@id='email']")
            self.type("//input[@id='email']", DataConfig.Email)
            self.wait_for_element("//input[@id='password']")
            self.type("//input[@id='password']", DataConfig.Password)
            self.click("//button[normalize-space()='Sign in']")
            self.wait_for_element("//a[normalize-space()='Logout']")

            # Go to customers
            self.wait_for_element("//a[normalize-space()='Customers']")
            self.click("//a[normalize-space()='Customers']")
            self.wait_for_element("(//td)[1]")

            # Excel operation
            ws = wb.active
            ws['A1'] = 'taxid'
            ws['B1'] = 'customertype'
            ws['C1'] = 'firstname'
            ws['D1'] = 'lastname'
            ws['E1'] = 'telephone'
            ws['F1'] = 'mobile'
            ws['G1'] = 'dob'
            ws['H1'] = 'pob'
            ws['I1'] = 'citizenship'
            ws['J1'] = 'street1'
            ws['K1'] = 'street2'
            ws['L1'] = 'city'
            ws['M1'] = 'region'
            ws['N1'] = 'postcode'

            pagination_no = DataConfig.PageNumber
            counter = 0

            for p in range(1, pagination_no + 1):
                self.click("//a[normalize-space()='%s']" % str(p))
                time.sleep(1)
                tax_id_lists = []

                for i in range(1, 21):
                    try:
                        list_element = self.get_text("//tbody/tr[%s]/td[1]" % str(i))
                        tax_id_lists.append(list_element)
                        ID = tax_id_lists[i - 1]
                        print(f"Now Collecting Client ID: {ID}")

                        self.open_new_tab()
                        self.get(f"https://app.cafcsn.it/#/pages/customers/edit/{tax_id_lists[i - 1]}")
                        time.sleep(2)

                        self.wait_for_element_visible("//input[@id='taxid']")
                        taxid = self.get_text("//input[@id='taxid']")

                        self.wait_for_element_clickable("//select[@id='customertype']")
                        s = Select(self.find_element("//select[@id='customertype']"))
                        customertype = s.first_selected_option.text

                        firstname = self.get_text("//input[@id='firstname']")
                        lastname = self.get_text("//input[@id='lastname']")
                        telephone = self.get_text("//input[@id='telephone']")
                        mobile = self.get_text("//input[@id='mobile']")
                        dob = self.get_text("//input[@id='dob']")
                        pob = self.get_text("//input[@id='pob']")
                        citizenship = self.get_text("//input[@id='citizenship']")
                        street1 = self.get_text("//input[@id='street1']")
                        street2 = self.get_text("//input[@id='street2']")
                        city = self.get_text("//input[@id='city']")
                        region = self.get_text("//input[@id='region']")
                        postcode = self.get_text("//input[@id='postcode']")

                        ws.append([taxid, customertype, firstname, lastname, telephone, mobile, dob, pob, citizenship,
                                   street1, street2, city, region, postcode])
                        counter += 1
                        self.post_message(f"Total Record Collected: {counter}")
                        print(f"Total Record Collected: {counter}")
                    except (NoSuchElementException, TimeoutException) as e:
                        print(f"An error occurred while processing client ID: {ID}. Error: {e}")
                    finally:
                        self.driver.close()
                        self.switch_to_default_tab()

                wb.save("datafile.xlsx")

        except Exception as e:
            print(f"An error occurred: {e}")
            wb.save("datafile.xlsx")
