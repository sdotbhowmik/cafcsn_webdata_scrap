import time
from seleniumbase import BaseCase
from openpyxl import load_workbook
from selenium.webdriver.support.select import Select
from config.config import DataConfig

class FetchDetails(BaseCase):

    def test_fetch_details(self):
        self.driver.maximize_window()
        self.driver.implicitly_wait(30)
        self.driver.set_page_load_timeout(30)

        try:
            self.open("https://app.cafcsn.it/#/login")
            self.wait_for_element("//input[@id='email']")
            self.type("//input[@id='email']", DataConfig.Email)
            self.wait_for_element("//input[@id='password']")
            self.type("//input[@id='password']", DataConfig.Password)
            self.click("//button[normalize-space()='Sign in']")
            self.wait_for_element("//a[normalize-space()='Logout']")

            # Load the Excel file with collected IDs
            wb_ids = load_workbook("client_ids.xlsx")
            ws_ids = wb_ids.active

            # Load or create the client_details.xlsx file
            try:
                wb_details = load_workbook("client_details.xlsx")
                ws_details = wb_details.active
            except FileNotFoundError:
                wb_details = Workbook()
                ws_details = wb_details.active
                # Add headers for the details if the file doesn't exist
                ws_details.append([
                    'customerid', 'taxid', 'customertype', 'firstname', 'lastname',
                    'telephone', 'mobile', 'dob', 'pob', 'citizenship',
                    'street1', 'street2', 'city', 'region', 'postcode'
                ])

            # Load the page range from config
            max_page = DataConfig.PageNumber

            # Find the first unprocessed row
            for row in range(2, ws_ids.max_row + 1):
                status = ws_ids.cell(row=row, column=3).value
                if status == "Pending":
                    customerid = ws_ids.cell(row=row, column=1).value
                    pagenumber = ws_ids.cell(row=row, column=2).value

                    if pagenumber > max_page:
                        print(f"Reached the maximum page limit of {max_page}. Stopping execution.")
                        break

                    try:
                        self.open_new_tab()
                        self.get(f"https://app.cafcsn.it/#/pages/customers/edit/{customerid}")
                        time.sleep(2)

                        self.wait_for_element_visible('//button[@class="btn btn-danger float-md-right"]')
                        taxid = self.get_text("//input[@id='taxid']")
                        customertype = Select(self.find_element("//select[@id='customertype']")).first_selected_option.text
                        firstname = self.get_text("//input[@id='firstname']")
                        lastname = self.get_text("//input[@id='lastname']")
                        telephone = self.get_text("//input[@id='telephone']")
                        mobile = self.get_text("//input[@id='mobile']")
                        dob = self.get_text("//input[@id='dob']")
                        pob = self.get_text("//input[@id='pob']")
                        citizenship = self.get_text("//input[@id='citizenship']")
                        street1 = self.get_text("//input[@id='street1']")
                        street2 = self.get_text("//input[@id='street2']")
                        city = self.get_text("//input[@id='city']")
                        region = self.get_text("//input[@id='region']")
                        postcode = self.get_text("//input[@id='postcode']")

                        # Update Excel sheet with details in client_details.xlsx
                        ws_details.append([
                            customerid, taxid, customertype, firstname, lastname,
                            telephone, mobile, dob, pob, citizenship,
                            street1, street2, city, region, postcode
                        ])

                        # Update the status to "Processed" in client_ids.xlsx
                        ws_ids.cell(row=row, column=3).value = "Processed"
                        print(f"Updated details for Client ID: {customerid} from Page: {pagenumber}")

                    except Exception as e:
                        print(f"An error occurred while fetching details for Client ID: {customerid}. Error: {e}")
                    finally:
                        self.driver.close()
                        self.switch_to_default_tab()

            # Save the updated Excel files
            wb_ids.save("client_ids.xlsx")
            wb_details.save("client_details.xlsx")

        except Exception as e:
            print(f"An error occurred during details fetching: {e}")
            wb_ids.save("client_ids.xlsx")
            wb_details.save("client_details.xlsx")